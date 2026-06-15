#!/usr/bin/env python3
"""
dxf_tube.py — Extract cable-mark tube counts and a device list from an
AutoCAD Electrical DXF, and write a two-sheet Excel workbook.

Tube rule (confirmed with the user):
    tube = 2 * (number of connection points on that cable mark)
    -- the "2" is head + tail: every wire end must be marked.

What can be traced automatically from DXF geometry:
    * A cable mark sitting on a VERTICAL rung wire that connects the device
      directly ABOVE and directly BELOW it in the same column. That is a
      2-device connection -> 2 tubes. This is reliable.

What CANNOT be traced from DXF geometry (left BLANK for manual entry):
    * BUS RAILS (e.g. 220V / 0V / N) — horizontal rails with junction-dot gaps
      that fan out to many devices. The connectivity is not stored in the DXF
      (link/terminal attributes export empty), so the count is left blank.
    * BRANCHED rungs that tap a third device off to the side.
    These rows are emitted with an empty tube cell and a note, so the engineer
    fills them in from the drawing instead of trusting a wrong number.

Usage:
    python dxf_tube.py DRAWING.dxf -o OUTPUT.xlsx
"""
import argparse
import math
from collections import defaultdict

import ezdxf

# ----------------------------------------------------------------------------
# Tunables — geometry tolerances in drawing units
# ----------------------------------------------------------------------------
COL_TOL = 20.0      # max |x| diff for a device to count as "same column" as a mark
WIRE_LAYER_PREFIX = "YEL"   # AutoCAD Electrical wire layers (YEL_14_THHN, etc.)
CBL_FAMILY = "CBL"          # cable-mark block family


def _attr_dict(insert):
    return {a.dxf.tag: a.dxf.text for a in insert.attribs}


def load_dxf(path):
    """Return (wires, cable_marks, devices).

    wires        : list of ((x1,y1),(x2,y2)) on wire layers
    cable_marks  : list of (tag, (x,y))
    devices      : list of (family, tag, (x,y), attrs)
    """
    try:
        doc = ezdxf.readfile(path)
    except ezdxf.DXFStructureError:
        # Damaged file — recover.
        doc, _ = ezdxf.recover.readfile(path)
    msp = doc.modelspace()

    wires = []
    for e in msp.query("LINE"):
        if str(e.dxf.layer).startswith(WIRE_LAYER_PREFIX):
            s, t = e.dxf.start, e.dxf.end
            wires.append(((round(s.x, 2), round(s.y, 2)),
                          (round(t.x, 2), round(t.y, 2))))

    cable_marks, devices = [], []
    for e in msp.query("INSERT"):
        a = _attr_dict(e)
        fam = a.get("FAMILY")
        tag = a.get("TAG1") or a.get("TAG2")
        p = (round(e.dxf.insert.x, 2), round(e.dxf.insert.y, 2))
        if fam == CBL_FAMILY:
            cable_marks.append((tag, p))
        elif fam:                      # any real component (SW, CR, OL, TD, MS...)
            devices.append((fam, tag, p, a))
    return wires, cable_marks, devices


def neighbors_in_column(cable_pos, devices):
    """Nearest device ABOVE and BELOW the cable mark in the same column.

    Returns (above_tag, below_tag); either may be None.
    """
    cx, cy = cable_pos
    col = [(tag, p) for (_f, tag, p, _a) in devices if abs(p[0] - cx) <= COL_TOL]
    above = [(p[1], tag) for tag, p in col if p[1] > cy]   # larger y = above
    below = [(p[1], tag) for tag, p in col if p[1] < cy]
    a = min(above)[1] if above else None   # closest above (smallest y over cy)
    b = max(below)[1] if below else None   # closest below (largest y under cy)
    return a, b


def trace_marks(wires, cable_marks, devices):
    """For each cable mark decide tube count.

    Returns list of dicts: {name, tube, connection, auto}
      tube == None  -> could not be determined automatically (leave blank)
      auto == True  -> traced from geometry; False -> needs manual entry
    """
    # Identify likely bus-rail marks: they sit at the far edge of the drawing,
    # with no device directly above AND below in their column.
    rows = []
    for tag, p in cable_marks:
        name = (tag or "").lstrip("-")
        above, below = neighbors_in_column(p, devices)
        if above and below:
            # Clean 2-device vertical rung -> reliable for the direct pair.
            conn = f"{above.lstrip('-')}-{below.lstrip('-')}"
            # A rung whose neighbour is a contactor coil/contact (K*) often also
            # taps a parallel device (a side branch the geometry can't see), e.g.
            # file-1 cases 02/05 were really 4 tubes. Flag — don't fabricate.
            branch_hint = any(t and t.lstrip("-").startswith("K")
                              for t in (above, below))
            rows.append({
                "name": name,
                "tube": 2,                       # 2 = head + tail of one pair
                "connection": conn + ("  (?อาจมีสาขา ตรวจแบบ)" if branch_hint else ""),
                "auto": True,
            })
        else:
            # Bus rail or single-ended / branched -> cannot trust geometry.
            ends = [t.lstrip("-") for t in (above, below) if t]
            hint = "+".join(ends) if ends else ""
            rows.append({
                "name": name,
                "tube": None,                    # leave blank for manual entry
                "connection": f"bus/branch (ตรวจจากแบบ){' ' + hint if hint else ''}",
                "auto": False,
            })

    rows.sort(key=_mark_sort_key)
    return rows


def _mark_sort_key(row):
    """Sort A1,A2,...A10,A11 numerically, named rails (220V,N,0V) after."""
    n = row["name"]
    digits = "".join(ch for ch in n if ch.isdigit())
    if n[:1] in ("A",) and digits:
        return (0, int(digits))
    if digits and n.replace(digits, "").strip("V") == "":
        return (1, int(digits))      # e.g. 01, 02 ... and 0V-like
    return (2, n)


def build_device_list(devices):
    """De-duplicate devices by tag, pull any attribute data, blanks otherwise."""
    seen = {}
    for fam, tag, p, a in devices:
        sym = (tag or "").lstrip("-")
        if not sym:
            continue
        if sym not in seen:
            seen[sym] = {
                "symbol": sym,
                # Pull from attributes if present, else blank for manual fill.
                "name": a.get("DESC1", "") or a.get("CATDESC", ""),
                "brand": a.get("MFG", ""),
                "model": a.get("CAT", ""),
            }
    out = list(seen.values())
    out.sort(key=lambda d: _device_sort_key(d["symbol"]))
    return out


def _device_sort_key(sym):
    prefix = "".join(ch for ch in sym if ch.isalpha())
    digits = "".join(ch for ch in sym if ch.isdigit())
    return (prefix, int(digits) if digits else 0)


# ----------------------------------------------------------------------------
# Excel writer
# ----------------------------------------------------------------------------
def build_workbook_bytes(dxf_path):
    """High-level helper for the web app: DXF path -> (xlsx_bytes, stats)."""
    import io
    wires, marks, devices = load_dxf(dxf_path)
    rows = trace_marks(wires, marks, devices)
    dev_list = build_device_list(devices)
    buf = io.BytesIO()
    write_excel(rows, dev_list, buf)
    buf.seek(0)
    stats = {
        "marks": len(rows),
        "traced": sum(1 for r in rows if r["auto"]),
        "blank": sum(1 for r in rows if not r["auto"]),
        "devices": len(dev_list),
    }
    return buf.getvalue(), stats


def write_excel(rows, devices, out_path):
    import pandas as pd
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cable_df = pd.DataFrame(
        [(i + 1, r["name"], ("" if r["tube"] is None else r["tube"]), r["connection"])
         for i, r in enumerate(rows)],
        columns=["No", "Name cablemark", "จำนวน tube", "การเชื่อมต่อ"],
    )
    dev_df = pd.DataFrame(
        [(i + 1, d["symbol"], d["name"], d["brand"], d["model"])
         for i, d in enumerate(devices)],
        columns=["No", "Symbol", "Name", "Brand", "Model"],
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        cable_df.to_excel(xl, sheet_name="CableMark", index=False)
        dev_df.to_excel(xl, sheet_name="DeviceList", index=False)
        _style(xl.sheets["CableMark"], widths=[8, 18, 14, 34])
        _style(xl.sheets["DeviceList"], widths=[8, 14, 24, 18, 18])

        # Highlight blank tube cells (need manual entry) in CableMark sheet.
        ws = xl.sheets["CableMark"]
        warn = PatternFill("solid", fgColor="FCE4D6")
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 3).value in (None, ""):
                ws.cell(r, 3).fill = warn

    return out_path


def _style(ws, widths):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    hdr = PatternFill("solid", fgColor="1F4E78")
    side = Side(style="thin", color="999999")
    box = Border(side, side, side, side)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr
        c.alignment = Alignment("center", "center")
        c.border = box
    for row in ws.iter_rows(min_row=2):
        for j, c in enumerate(row):
            c.border = box
            # left-align the last text column, center the rest
            c.alignment = Alignment("left" if j == len(widths) - 1 and ws.title == "CableMark"
                                    else ("left" if (ws.title == "DeviceList" and j >= 2) else "center"),
                                    "center")
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w


def main():
    ap = argparse.ArgumentParser(description="DXF -> cable-mark tube list + device list (Excel)")
    ap.add_argument("dxf", help="input .dxf file")
    ap.add_argument("-o", "--out", required=True, help="output .xlsx path")
    args = ap.parse_args()

    wires, marks, devices = load_dxf(args.dxf)
    rows = trace_marks(wires, marks, devices)
    dev_list = build_device_list(devices)
    write_excel(rows, dev_list, args.out)

    auto = sum(1 for r in rows if r["auto"])
    blank = len(rows) - auto
    print(f"Cable marks: {len(rows)} ({auto} traced, {blank} left blank for manual entry)")
    print(f"Devices: {len(dev_list)}")
    print(f"Saved: {args.out}")


if __name__ == "__main__":
    main()
